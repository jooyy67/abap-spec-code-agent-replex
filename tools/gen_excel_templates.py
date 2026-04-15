from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="E2E8F0")
    font = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _set_widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def create_selection_template(out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "selection"

    headers = [
        "order",
        "section",
        "label",
        "field_id",
        "input_type",
        "is_required",
        "has_f4",
        "ui_rule",
    ]
    ws.append(headers)

    # Sample rows (ASCII-safe; Korean values are entered by users)
    sample = [
        ["1", "Basic", "Company Code", "BUKRS", "단일", "Y", "Y", "1000 fixed"],
        ["2", "Basic", "Posting Date", "BUDAT", "범위", "N", "N", ""],
        ["3", "Option", "Plant", "WERKS", "리스트", "N", "Y", ""],
    ]
    for r in sample:
        ws.append(r)

    _style_header(ws)
    _set_widths(
        ws,
        {
            "A": 8,
            "B": 14,
            "C": 18,
            "D": 14,
            "E": 12,
            "F": 12,
            "G": 10,
            "H": 22,
        },
    )
    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def create_alv_template(out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "alv"

    headers = [
        "area",  # A/B (optional). If omitted, defaults to A.
        "order",
        "label",
        "field_id",
        "is_key",
        "is_sum",
        "is_edit",
        "action",
    ]
    ws.append(headers)

    sample = [
        ["A", "1", "Document No", "BELNR", "Y", "N", "N", "hotspot click → detail"],
        ["A", "2", "Status", "STATUS", "N", "N", "N", "icon"],
        ["B", "1", "Item", "BUZEI", "Y", "N", "N", ""],
        ["B", "2", "Qty", "MENGE", "N", "Y", "Y", "data_changed recalculation"],
    ]
    for r in sample:
        ws.append(r)

    _style_header(ws)
    _set_widths(
        ws,
        {
            "A": 8,
            "B": 8,
            "C": 18,
            "D": 14,
            "E": 10,
            "F": 10,
            "G": 10,
            "H": 26,
        },
    )
    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    create_selection_template(root / "assets" / "templates" / "selection_template.xlsx")
    create_alv_template(root / "assets" / "templates" / "alv_template.xlsx")
    print("generated:", "assets/templates/selection_template.xlsx")
    print("generated:", "assets/templates/alv_template.xlsx")


if __name__ == "__main__":
    main()

