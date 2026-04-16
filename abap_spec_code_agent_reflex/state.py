import reflex as rx
from typing import List
from pydantic import BaseModel, Field  # 추가
import os
from pathlib import Path
import base64
import io
import json
import uuid
import re
import unicodedata
from datetime import datetime, timezone

import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import load_workbook
from PIL import Image

# 프로젝트 루트의 .env (gitignore됨). Windows 환경 변수가 있으면 그쪽이 우선.
load_dotenv(Path(__file__).resolve().parent.parent / ".env")


class SelectionField(BaseModel):
    order: str = ""
    section: str = ""
    label: str = ""
    field_id: str = ""
    input_type: str = "단일"
    is_required: bool = False
    has_f4: bool = False
    ui_rule: str = ""


class ALVField(BaseModel):
    order: str = ""
    label: str = ""
    field_id: str = ""
    is_key: bool = False
    is_sum: bool = False
    is_edit: bool = False
    action: str = ""
    # 화면 요약용(선택 행). 영역 A/B 각각 독립적으로 관리 (FS·JSON 입력 데이터에는 포함하지 않음).
    preview_pick_a: bool = False
    preview_pick_b: bool = False


class PickedAlvField(BaseModel):
    """업로드된 모든 테이블에서 체크된 필드 요약용."""

    table_label: str = ""
    order: str = ""
    label: str = ""
    field_id: str = ""
    is_key: bool = False
    is_sum: bool = False
    is_edit: bool = False
    action: str = ""


class UploadedAlvTable(BaseModel):
    """이미지 1장에서 추출된 ALV 필드 스냅샷 (업로드 테이블 목록 1행)."""

    id: str = ""
    label: str = ""
    fields: List[ALVField] = Field(default_factory=list)


class FsBlock(BaseModel):
    """FS 마크다운을 섹션 단위로 분리한 렌더링 블록."""

    title: str = ""
    content_md: str = ""


# --- 메인 상태 (내비게이션 관리) ---
class StateMain(rx.State):
    input_sections: List[str] = [
        "1. 기본 정보",
        "2. 프로그램 유형 & 상세 기능",
        "3. 화면 구조 및 동작",
        "4. 조회 조건 (Selection Screen)",
        "5. 조회 결과 리스트 (ALV Grid)",
        "6. 업무 처리 및 데이터 규칙 (Processing Logic)",
    ]
    input_section_ids: List[str] = [
        "input-1",
        "input-2",
        "input-3",
        "input-4",
        "input-5",
        "input-6",
    ]

    # FS 탭은 한 화면에서 (1) Functional Spec + (2) 개발 매핑 입력서를 2열로 함께 보여줍니다.
    fs_sections: List[str] = ["Functional Spec"]
    fs_section_ids: List[str] = ["fs-1"]

    code_sections: List[str] = ["Code 생성"]
    code_section_ids: List[str] = ["code-1"]

    sections: List[str] = input_sections + fs_sections + code_sections

    current_section: str = sections[0]
    hovered_section: str = ""
    expanded_sidebar_group: str = "input"
    sidebar_collapsed: bool = False

    def set_current_section(self, section: str):
        self.current_section = section

    def set_hovered_section(self, section: str):
        self.hovered_section = section

    def toggle_sidebar_group(self, group: str):
        self.expanded_sidebar_group = "" if self.expanded_sidebar_group == group else group

    @rx.event
    def toggle_sidebar(self):
        self.sidebar_collapsed = not bool(self.sidebar_collapsed)

    def toggle_input_group(self):
        was_open = self.expanded_sidebar_group == "input"
        self.toggle_sidebar_group("input")
        if not was_open:
            self.current_section = self.input_sections[0]

    def toggle_fs_group(self):
        was_open = self.expanded_sidebar_group == "fs"
        self.toggle_sidebar_group("fs")
        if not was_open:
            self.current_section = self.fs_sections[0]

    def toggle_code_group(self):
        was_open = self.expanded_sidebar_group == "code"
        self.toggle_sidebar_group("code")
        if not was_open:
            self.current_section = self.code_sections[0]

    @staticmethod
    def _split_md_sections(md: str) -> dict[str, str]:
        """
        Backward-compatible helper.
        Split markdown by '## ' headings.
        Returns {heading_title: block_including_heading_line}.
        """
        text = (md or "").replace("\r\n", "\n")
        parts = re.split(r"(?m)^##\s+", text)
        out: dict[str, str] = {}
        if not parts:
            return out
        # parts[0] is preface (e.g. '# ...')
        for p in parts[1:]:
            lines = p.split("\n")
            title = (lines[0] or "").strip()
            body = "\n".join(lines[1:]).strip("\n")
            out[title] = f"## {title}\n{body}".strip() + "\n"
        return out


# --- 데이터 상태 (입력값 관리) ---
class State(StateMain):  # StateMain을 상속받아 하나로 합침 (중복 방지)
    # 섹션 1
    prog_name: str = ""
    req_dept_user: str = ""
    work_category: str = ""
    ref_tcode: str = ""
    auth_range: List[str] = []
    work_options: List[str] = [
        "구매(MM)",
        "영업(SD)",
        "생산(PP)",
        "회계(FI)",
        "인사(HR)",
        "원가(CO)",
        "자산(AA)",
        "기타",
    ]

    # 섹션 2
    prog_type: str = "조회 전용"
    selected_features: List[str] = []

    # 섹션 3
    ui_layout: str = "단일 화면"
    layout_detail: str = ""
    ui_features: List[str] = []
    edit_auth: str = "조회 전용"
    other_ui_notes: str = ""
    # 3.x 신규 UI 상태(요청된 3번 구조 반영)
    grid_count: str = "단일 Grid"  # 단일 Grid / 다중 Grid
    split_mode: str = "상하"  # 상하 / 좌우 / 탭
    click_actions: List[str] = []  # 없음 / 화면 이동 / 팝업 / 내부연동 (Drill-Down)
    exec_methods: List[str] = []  # Hotspot / Double Click / 버튼
    detail_a: List[str] = []  # 단순 조회 / 상세 조회 사용
    detail_b_scope: List[str] = []  # 신규 추가 / 수정 가능 / 삭제 가능
    detail_b_input_methods: List[str] = []  # 엑셀 업로드 / 직접 입력 / 팝업 / 화면이동
    detail_b_single_grid_controls: List[str] = []  # 행 추가 / 행 삭제 / 모드 전환 / 다건 처리
    detail_b_multi_top: List[str] = []  # 조회 전용 / 편집 가능(추가/삭제/모드전환)
    detail_b_multi_bottom: List[str] = []  # 조회 전용 / 편집 가능(추가/삭제/모드전환/다건처리)

    # 섹션 4
    selection_fields: List[SelectionField] = [
        SelectionField(
            order="1",
            section="기본",
            label="회사 코드",
            field_id="BUKRS",
            input_type="단일",
            is_required=True,
            has_f4=True,
            ui_rule="1000 고정",
        )
    ]

    # 섹션 5
    alv_fields_a: List[ALVField] = [
        ALVField(
            order="1",
            label="전표 번호",
            field_id="BELNR",
            is_key=True,
            action="클릭 시 상세 조회",
        ),
        ALVField(order="2", label="처리 상태", field_id="STATUS", action="아이콘 표시"),
    ]
    alv_fields_b: List[ALVField] = [
        ALVField(order="1", label="품목 번호", field_id="BUZEI", is_key=True),
        ALVField(
            order="2",
            label="수량",
            field_id="MENGE",
            is_sum=True,
            is_edit=True,
            action="값 변경 시 재계산",
        ),
    ]
    alv_picked_preview_a: List[ALVField] = []
    alv_picked_preview_b: List[ALVField] = []
    alv_picked_preview_all_a: List[PickedAlvField] = []
    alv_picked_preview_all_b: List[PickedAlvField] = []

    # 섹션 6
    logic_flow: str = ""
    validation_rule: str = ""
    auto_calc_rule: str = ""
    data_target: List[str] = []
    delete_method: str = "미지정"
    dup_check: str = "미지정"

    generated_fs: str = ""
    fs_error: str = ""
    is_generating_fs: bool = False
    fs_blocks: List[FsBlock] = []
    generated_mapping_spec: str = ""
    mapping_error: str = ""
    mapping_sections: dict[str, str] = {}
    mapping_updated_at: str = ""
    mapping_active_section: str = "기능 매핑"
    mapping_edit_section: str = ""

    # 입력 보완(팝업) 상태: 원본 데이터 편집 → FS/매핑 재생성
    edit_modal_open: bool = False
    edit_field_key: str = "prog_name"
    edit_field_value: str = ""
    edit_modal_error: str = ""

    # --- 업로드(자동 입력) 상태 ---
    # NOTE: 4번(Selection) / 5번(ALV) 업로드는 서로 다른 파일/흐름이므로 상태를 분리.
    selection_upload_error: str = ""
    selection_upload_info: str = ""
    selection_upload_debug: str = ""
    show_selection_upload_debug: bool = False

    alv_upload_error: str = ""
    alv_upload_info: str = ""
    alv_upload_debug: str = ""
    show_alv_upload_debug: bool = False
    image_upload_error: str = ""
    image_upload_info: str = ""
    # SE11 이미지 → 테이블 목록 (이미지 1장당 테이블 1개, 병합 없음)
    se11_paste_queue_uris: List[str] = []
    uploaded_alv_tables: List[UploadedAlvTable] = []
    se11_select_options: List[str] = ["(선택 없음)"]
    area_bind_value_a: str = "(선택 없음)"
    area_bind_value_b: str = "(선택 없음)"
    area_bind_id_a: str = ""
    area_bind_id_b: str = ""

    @rx.event
    def set_show_selection_upload_debug(self, value: bool):
        self.show_selection_upload_debug = bool(value)

    @rx.event
    def set_show_alv_upload_debug(self, value: bool):
        self.show_alv_upload_debug = bool(value)

    @rx.event
    def clear_selection_upload_and_table(self):
        """4번(조회조건) 선택 취소: 선택된 파일 해제 + 표 초기화 + 메시지/디버그 초기화."""
        self.selection_fields = [SelectionField(order="1")]
        self.selection_upload_error = ""
        self.selection_upload_info = ""
        self.selection_upload_debug = ""
        self.show_selection_upload_debug = False

    @rx.event
    def clear_alv_upload_and_tables(self):
        """5번(ALV) 선택 취소: 선택된 파일 해제 + 표 초기화 + 메시지/디버그 초기화."""
        self.alv_fields_a = [ALVField(order="1")]
        self.alv_fields_b = [ALVField(order="1")]
        self.alv_picked_preview_a = []
        self.alv_picked_preview_b = []
        self.alv_picked_preview_all_a = []
        self.alv_picked_preview_all_b = []
        self.alv_upload_error = ""
        self.alv_upload_info = ""
        self.alv_upload_debug = ""
        self.show_alv_upload_debug = False
        self._reset_se11_uploaded_data()

    def _refresh_fs_blocks(self):
        """generated_fs를 섹션(## ...) 기준으로 분리해 fs_blocks를 갱신."""
        md = (self.generated_fs or "").replace("\r\n", "\n").strip()
        if not md:
            self.fs_blocks = []
            return

        blocks: list[FsBlock] = []
        curr_title: str | None = None
        curr_lines: list[str] = []

        for line in md.split("\n"):
            m = re.match(r"^##\s+(.+?)\s*$", line)
            if m:
                # flush previous
                if curr_title is not None:
                    content = "\n".join(curr_lines).strip()
                    blocks.append(FsBlock(title=curr_title, content_md=content))
                curr_title = m.group(1).strip()
                curr_lines = []
                continue
            curr_lines.append(line)

        if curr_title is not None:
            content = "\n".join(curr_lines).strip()
            blocks.append(FsBlock(title=curr_title, content_md=content))
        else:
            # No "##" sections found: render whole doc as one block.
            blocks = [FsBlock(title="Functional Spec", content_md=md)]

        self.fs_blocks = blocks

    def _reset_se11_uploaded_data(self):
        self.se11_paste_queue_uris = []
        self.uploaded_alv_tables = []
        self._sync_se11_select_options()
        self.area_bind_value_a = "(선택 없음)"
        self.area_bind_value_b = "(선택 없음)"
        self.area_bind_id_a = ""
        self.area_bind_id_b = ""
        self.image_upload_error = ""
        self.image_upload_info = ""

    @rx.event
    def reset_se11_tables_and_bindings(self):
        """업로드 테이블 목록·바인딩·영역 A/B 표를 초기화합니다."""
        self._reset_se11_uploaded_data()
        self.alv_fields_a = [ALVField(order="1")]
        self.alv_fields_b = [ALVField(order="1")]
        self.alv_picked_preview_a = []
        self.alv_picked_preview_b = []
        self.alv_picked_preview_all_a = []
        self.alv_picked_preview_all_b = []

    @rx.event
    def remove_se11_paste_queue_item(self, index: int):
        """Ctrl+V로 붙여넣은 이미지 대기열에서 특정 이미지를 제거."""
        try:
            idx = int(index)
        except Exception:
            return
        if idx < 0 or idx >= len(self.se11_paste_queue_uris or []):
            return
        q = list(self.se11_paste_queue_uris or [])
        q.pop(idx)
        self.se11_paste_queue_uris = q

    @rx.event
    def delete_uploaded_alv_table(self, table_id: str):
        """생성된 업로드 테이블(라이브러리)에서 특정 테이블을 삭제."""
        tid = (table_id or "").strip()
        if not tid:
            return

        self.uploaded_alv_tables = [t for t in (self.uploaded_alv_tables or []) if t.id != tid]
        self._sync_se11_select_options()

        # 바인딩이 삭제된 테이블을 가리키면 해제하고 표를 초기화.
        if self.area_bind_id_a == tid:
            self.area_bind_id_a = ""
            self.area_bind_value_a = "(선택 없음)"
            self.alv_fields_a = [self._empty_alv_row()]
            self._refresh_alv_picked_preview("A")
        if self.area_bind_id_b == tid:
            self.area_bind_id_b = ""
            self.area_bind_value_b = "(선택 없음)"
            self.alv_fields_b = [self._empty_alv_row()]
            self._refresh_alv_picked_preview("B")

        # 전체 요약도 다시 계산
        self._refresh_alv_picked_preview_all("A")
        self._refresh_alv_picked_preview_all("B")

    # --- 핸들러 함수들 ---
    def set_prog_name(self, value: str):
        self.prog_name = value

    def set_req_dept_user(self, value: str):
        self.req_dept_user = value

    def set_work_category(self, value: str):
        self.work_category = value

    def set_prog_type(self, value: str):
        self.prog_type = value

    def set_ui_layout(self, value: str):
        self.ui_layout = value

    def set_layout_detail(self, value: str):
        self.layout_detail = value

    def set_edit_auth(self, value: str):
        self.edit_auth = value

    def set_other_ui_notes(self, value: str):
        self.other_ui_notes = value

    def set_grid_count(self, value: str):
        self.grid_count = value

    def set_split_mode(self, value: str):
        self.split_mode = value

    def toggle_detail_b_scope(self, value: str, checked: bool):
        """3.4-B-1 처리 범위 토글. '신규 추가' 해제 시 입력방식도 초기화."""
        curr = getattr(self, "detail_b_scope", []).copy()
        if checked:
            if value not in curr:
                curr.append(value)
        else:
            if value in curr:
                curr.remove(value)
            if value == "신규 추가":
                self.detail_b_input_methods = []
        self.detail_b_scope = curr

    def update_auth_range(self, value: str, checked: bool):
        if checked:
            self.auth_range.append(value)
        else:
            self.auth_range.remove(value)

    def add_selection_row(self):
        self.selection_fields.append(
            SelectionField(order=str(len(self.selection_fields) + 1))
        )

    def update_field(self, index: int, key: str, value: any):
        setattr(self.selection_fields[index], key, value)

    def delete_selection_row(self, index: int):
        self.selection_fields.pop(index)

    def update_list(self, attr_name: str, value: str, checked: bool):
        curr = getattr(self, attr_name).copy()
        if checked:
            curr.append(value)
        else:
            curr.remove(value)
        setattr(self, attr_name, curr)

    def _sync_se11_select_options(self):
        self.se11_select_options = ["(선택 없음)"] + [t.label for t in (self.uploaded_alv_tables or [])]

    def _empty_alv_row(self) -> ALVField:
        return ALVField(order="1")

    def _resolve_bind_id_from_label(self, label: str) -> str:
        if not label or label == "(선택 없음)":
            return ""
        for t in self.uploaded_alv_tables or []:
            if t.label == label:
                return t.id
        return ""

    def _display_from_uploaded_label(self, area: str, label: str):
        tid = self._resolve_bind_id_from_label(label)
        if area == "A":
            self.area_bind_value_a = label
            self.area_bind_id_a = tid
        else:
            self.area_bind_value_b = label
            self.area_bind_id_b = tid
        if not tid:
            if area == "A":
                self.alv_fields_a = [self._empty_alv_row()]
                self._refresh_alv_picked_preview("A")
            else:
                self.alv_fields_b = [self._empty_alv_row()]
                self._refresh_alv_picked_preview("B")
            return
        for t in self.uploaded_alv_tables or []:
            if t.id == tid:
                rows = [f.model_copy(deep=True) for f in t.fields] if t.fields else [self._empty_alv_row()]
                if area == "A":
                    self.alv_fields_a = rows
                else:
                    self.alv_fields_b = rows
                self._refresh_alv_picked_preview(area)
                return
        if area == "A":
            self.alv_fields_a = [self._empty_alv_row()]
            self._refresh_alv_picked_preview("A")
        else:
            self.alv_fields_b = [self._empty_alv_row()]
            self._refresh_alv_picked_preview("B")

    def _refresh_alv_picked_preview(self, area: str):
        fields = self.alv_fields_a if area == "A" else self.alv_fields_b
        key = "preview_pick_a" if area == "A" else "preview_pick_b"
        sub = [f.model_copy(deep=True) for f in fields if getattr(f, key, False)]
        if area == "A":
            self.alv_picked_preview_a = sub
        else:
            self.alv_picked_preview_b = sub
        self._refresh_alv_picked_preview_all(area)

    def _refresh_alv_picked_preview_all(self, area: str):
        """업로드된 전체 테이블에서 체크된 필드를 모아 요약 리스트를 갱신."""
        picked: list[PickedAlvField] = []
        key = "preview_pick_a" if area == "A" else "preview_pick_b"
        for t in self.uploaded_alv_tables or []:
            for f in t.fields or []:
                if getattr(f, key, False):
                    picked.append(
                        PickedAlvField(
                            table_label=t.label,
                            order=f.order,
                            label=f.label,
                            field_id=f.field_id,
                            is_key=f.is_key,
                            is_sum=f.is_sum,
                            is_edit=f.is_edit,
                            action=f.action,
                        )
                    )
        if area == "A":
            self.alv_picked_preview_all_a = picked
        else:
            self.alv_picked_preview_all_b = picked

    @rx.event
    def set_area_bind_a(self, value: str):
        self._display_from_uploaded_label("A", value)

    @rx.event
    def set_area_bind_b(self, value: str):
        self._display_from_uploaded_label("B", value)

    def _unique_se11_table_label(self, suggested: str) -> str:
        raw = self._clean_str(suggested or "").strip()
        base = raw if raw else f"업로드 테이블 {len(self.uploaded_alv_tables) + 1}"
        existing = {t.label for t in (self.uploaded_alv_tables or [])}
        if base not in existing:
            return base
        i = 2
        while f"{base} ({i})" in existing:
            i += 1
        return f"{base} ({i})"

    def _push_se11_display_to_library(self, area: str):
        tid = self.area_bind_id_a if area == "A" else self.area_bind_id_b
        if not tid:
            return
        src = self.alv_fields_a if area == "A" else self.alv_fields_b
        new_tables: list[UploadedAlvTable] = []
        for t in self.uploaded_alv_tables or []:
            if t.id == tid:
                new_tables.append(
                    UploadedAlvTable(
                        id=t.id,
                        label=t.label,
                        fields=[x.model_copy(deep=True) for x in src],
                    )
                )
            else:
                new_tables.append(t.model_copy(deep=True))
        self.uploaded_alv_tables = new_tables
        other = "B" if area == "A" else "A"
        other_id = self.area_bind_id_b if area == "A" else self.area_bind_id_a
        if other_id and other_id == tid:
            rows = [x.model_copy(deep=True) for x in src]
            if other == "A":
                self.alv_fields_a = rows
            else:
                self.alv_fields_b = rows
            self._refresh_alv_picked_preview(other)

    def add_alv_row(self, area: str):
        target = self.alv_fields_a if area == "A" else self.alv_fields_b
        target.append(ALVField(order=str(len(target) + 1)))
        self._push_se11_display_to_library(area)
        self._refresh_alv_picked_preview(area)

    def update_alv_field(self, area: str, index: int, key: str, value: any):
        target = self.alv_fields_a if area == "A" else self.alv_fields_b
        setattr(target[index], key, value)
        self._push_se11_display_to_library(area)
        self._refresh_alv_picked_preview(area)

    def delete_alv_row(self, area: str, index: int):
        target = self.alv_fields_a if area == "A" else self.alv_fields_b
        target.pop(index)
        self._push_se11_display_to_library(area)
        self._refresh_alv_picked_preview(area)

    def _clear_se11_bindings_on_excel_fill(self):
        """엑셀이 ALV를 직접 덮어쓸 때 이미지 바인딩을 해제합니다."""
        self.area_bind_value_a = "(선택 없음)"
        self.area_bind_value_b = "(선택 없음)"
        self.area_bind_id_a = ""
        self.area_bind_id_b = ""

    def handle_submit(self):
        if not self.prog_name:
            return rx.window_alert("프로그램 이름을 입력해주세요!")
        return self.generate_fs()

    # -----------------------------
    # 업로드 유틸 / 매핑 정의
    # -----------------------------
    @staticmethod
    def _norm_header(v: object) -> str:
        s = (str(v) if v is not None else "").strip().lower()
        s = re.sub(r"\s+", "", s)
        s = s.replace("-", "").replace("_", "")
        return s

    @staticmethod
    def _to_bool(v: object) -> bool:
        if isinstance(v, bool):
            return v
        if v is None:
            return False
        # Excel may contain full-width characters like 'Ｙ/Ｎ'. Normalize to improve robustness.
        s = unicodedata.normalize("NFKC", str(v)).strip().lower()
        return s in {"y", "yes", "true", "1", "t", "o", "ㅇ", "예", "필수"}

    @staticmethod
    def _clean_str(v: object) -> str:
        return "" if v is None else str(v).strip()

    @staticmethod
    def _extract_json(text: str) -> object:
        """
        Gemini가 주변 텍스트를 섞어도 JSON만 최대한 추출.
        """
        t = (text or "").strip()
        if not t:
            raise ValueError("empty response")

        # Prefer fenced json block if present.
        fence = re.search(r"```json\s*([\s\S]*?)```", t, re.IGNORECASE)
        if fence:
            t = fence.group(1).strip()

        # Try direct loads.
        try:
            return json.loads(t)
        except Exception:
            pass

        # Heuristic: first {...} or [...]
        m = re.search(r"(\[.*\]|\{.*\})", t, re.DOTALL)
        if not m:
            raise ValueError("no json found in response")
        return json.loads(m.group(1))

    def _apply_selection_rows(self, rows: list[dict], mode: str = "replace") -> int:
        parsed: list[SelectionField] = []
        for i, r in enumerate(rows):
            if not any(str(v).strip() for v in r.values() if v is not None):
                continue

            # Be lenient about key variants (some loaders may not map headers as expected).
            is_required_val = (
                r.get("is_required")
                if "is_required" in r
                else r.get("isrequired", r.get("required", r.get("필수")))
            )
            has_f4_val = (
                r.get("has_f4")
                if "has_f4" in r
                else r.get("hasf4", r.get("f4", r.get("F4")))
            )
            parsed.append(
                SelectionField(
                    order=self._clean_str(r.get("order")) or str(i + 1),
                    section=self._clean_str(r.get("section")),
                    label=self._clean_str(r.get("label")),
                    field_id=self._clean_str(r.get("field_id")),
                    input_type=self._clean_str(r.get("input_type")) or "단일",
                    is_required=self._to_bool(is_required_val),
                    has_f4=self._to_bool(has_f4_val),
                    ui_rule=self._clean_str(r.get("ui_rule")),
                )
            )

        if not parsed:
            return 0

        if mode == "append":
            self.selection_fields = (self.selection_fields or []) + parsed
        else:
            self.selection_fields = parsed
        return len(parsed)

    def _apply_alv_rows(self, rows: list[dict], area: str, mode: str = "replace") -> int:
        parsed: list[ALVField] = []
        for i, r in enumerate(rows):
            if not any(str(v).strip() for v in r.values() if v is not None):
                continue
            parsed.append(
                ALVField(
                    order=self._clean_str(r.get("order")) or str(i + 1),
                    label=self._clean_str(r.get("label")),
                    field_id=self._clean_str(r.get("field_id")),
                    is_key=self._to_bool(r.get("is_key")),
                    is_sum=self._to_bool(r.get("is_sum")),
                    is_edit=self._to_bool(r.get("is_edit")),
                    action=self._clean_str(r.get("action")),
                )
            )

        if not parsed:
            return 0

        if area == "B":
            self.alv_fields_b = (self.alv_fields_b or []) + parsed if mode == "append" else parsed
        else:
            self.alv_fields_a = (self.alv_fields_a or []) + parsed if mode == "append" else parsed
        self._refresh_alv_picked_preview(area)
        return len(parsed)

    def _excel_rows_from_bytes(self, data: bytes) -> tuple[str, list[dict]]:
        wb = load_workbook(io.BytesIO(data), data_only=True)

        # Prefer named sheets if present.
        sheet = None
        for name in wb.sheetnames:
            n = name.strip().lower()
            if n in {"selection", "selection_fields", "조회조건", "조회조건정의", "4", "section4"}:
                sheet = wb[name]
                return ("selection", self._rows_from_sheet(sheet))
        for name in wb.sheetnames:
            n = name.strip().lower()
            if n in {"alv", "alv_fields", "조회결과", "조회결과리스트", "5", "section5"}:
                sheet = wb[name]
                return ("alv", self._rows_from_sheet(sheet))
        # Default: active sheet with auto-detect
        sheet = wb.active
        rows = self._rows_from_sheet(sheet)
        # Detect type by headers
        headers = {self._norm_header(h) for h in (rows[0].keys() if rows else [])}
        # Selection has at least one of these distinguishing columns.
        if {"section", "input_type", "inputtype", "ui_rule", "uirule", "has_f4", "hasf4", "is_required", "isrequired"} & headers:
            return ("selection", rows)
        return ("alv", rows)

    def _rows_from_sheet(self, sheet) -> list[dict]:
        values = list(sheet.values)
        if not values:
            return []
        header_row = values[0]
        headers = [self._norm_header(h) for h in header_row]

        def col_key(h: str) -> str | None:
            # Selection
            if h in {"order", "순서", "no"}:
                return "order"
            if h in {"section", "섹션", "섹션명", "group"}:
                return "section"
            if h in {"label", "항목명", "항목", "명칭", "설명"}:
                return "label"
            if h in {"fieldid", "field", "필드", "필드id", "필드명"}:
                return "field_id"
            if h in {"inputtype", "입력방식", "입력", "type"}:
                return "input_type"
            if h in {"isrequired", "required", "필수"}:
                return "is_required"
            if h in {"hasf4", "f4"}:
                return "has_f4"
            if h in {"uirule", "rule", "화면제어규칙", "제어규칙"}:
                return "ui_rule"

            # ALV
            if h in {"iskey", "key", "중요키", "k"}:
                return "is_key"
            if h in {"issum", "sum", "합계"}:
                return "is_sum"
            if h in {"isedit", "edit", "수정"}:
                return "is_edit"
            if h in {"action", "동작"}:
                return "action"
            if h in {"area", "영역"}:
                return "area"
            return None

        mapped_cols: list[str | None] = [col_key(h) for h in headers]
        out: list[dict] = []
        for row in values[1:]:
            if row is None:
                continue
            d: dict = {}
            for idx, cell in enumerate(row):
                key = mapped_cols[idx] if idx < len(mapped_cols) else None
                if not key:
                    continue
                d[key] = cell
            # skip fully empty
            if not any((str(v).strip() if v is not None else "") for v in d.values()):
                continue
            out.append(d)
        return out

    # -----------------------------
    # 엑셀 업로드 → 자동 입력
    # -----------------------------
    @rx.event
    async def handle_excel_upload(self, files: list[rx.UploadFile]):
        # legacy entrypoint: keep behavior but write into selection fields by autodetect.
        self.selection_upload_error = ""
        self.selection_upload_info = ""
        self.selection_upload_debug = ""
        self.alv_upload_error = ""
        self.alv_upload_info = ""
        self.alv_upload_debug = ""

        if not files:
            self.selection_upload_error = "업로드할 엑셀 파일을 선택해주세요."
            return

        f = files[0]
        try:
            data = await f.read()
            kind, rows = self._excel_rows_from_bytes(data)
            if not rows:
                raise ValueError("엑셀에서 데이터 행을 찾지 못했습니다. (첫 행은 헤더여야 합니다)")
            # debug (first 3 rows)
            try:
                dbg = json.dumps(
                    {"kind": kind, "sample_rows": rows[:3]},
                    ensure_ascii=False,
                    default=str,
                    indent=2,
                )
                if kind == "selection":
                    self.selection_upload_debug = dbg
                else:
                    self.alv_upload_debug = dbg
            except Exception:
                if kind == "selection":
                    self.selection_upload_debug = f"(debug serialization failed) kind={kind} rows={len(rows)}"
                else:
                    self.alv_upload_debug = f"(debug serialization failed) kind={kind} rows={len(rows)}"

            if kind == "selection":
                n = self._apply_selection_rows(rows, mode="replace")
                self.selection_upload_info = f"조회조건 테이블에 {n}건을 자동 반영했습니다. (필요 시 직접 수정 가능)"
            else:
                # If area column exists, split.
                has_area = any("area" in r for r in rows)
                if has_area:
                    a_rows = [r for r in rows if str(r.get("area") or "").strip().upper() != "B"]
                    b_rows = [r for r in rows if str(r.get("area") or "").strip().upper() == "B"]
                    na = self._apply_alv_rows(a_rows, area="A", mode="replace") if a_rows else 0
                    nb = self._apply_alv_rows(b_rows, area="B", mode="replace") if b_rows else 0
                    self._clear_se11_bindings_on_excel_fill()
                    self.alv_upload_info = f"ALV 테이블 자동 반영 완료: 영역A {na}건, 영역B {nb}건"
                else:
                    # Default target: A
                    na = self._apply_alv_rows(rows, area="A", mode="replace")
                    self._clear_se11_bindings_on_excel_fill()
                    self.alv_upload_info = f"ALV(영역 A) 테이블에 {na}건을 자동 반영했습니다. (영역 B는 수동 또는 area 컬럼 사용)"
        except Exception as e:
            self.selection_upload_error = f"엑셀 자동 입력 실패: {e}"

    async def _handle_excel_upload_target(self, files: list[rx.UploadFile], target: str):
        """
        target: 'selection' | 'alv'
        """
        if target == "selection":
            self.selection_upload_error = ""
            self.selection_upload_info = ""
            self.selection_upload_debug = ""
        else:
            self.alv_upload_error = ""
            self.alv_upload_info = ""
            self.alv_upload_debug = ""

        if not files:
            if target == "selection":
                self.selection_upload_error = "업로드할 엑셀 파일을 선택해주세요."
            else:
                self.alv_upload_error = "업로드할 엑셀 파일을 선택해주세요."
            return

        f = files[0]
        try:
            data = await f.read()
            kind, rows = self._excel_rows_from_bytes(data)
            if not rows:
                raise ValueError("엑셀에서 데이터 행을 찾지 못했습니다. (첫 행은 헤더여야 합니다)")
            # debug (first 3 rows)
            try:
                dbg = json.dumps(
                    {"detected_kind": kind, "forced_target": target, "sample_rows": rows[:3]},
                    ensure_ascii=False,
                    default=str,
                    indent=2,
                )
                if target == "selection":
                    self.selection_upload_debug = dbg
                else:
                    self.alv_upload_debug = dbg
            except Exception:
                if target == "selection":
                    self.selection_upload_debug = f"(debug serialization failed) kind={kind} target={target} rows={len(rows)}"
                else:
                    self.alv_upload_debug = f"(debug serialization failed) kind={kind} target={target} rows={len(rows)}"

            if target == "selection":
                n = self._apply_selection_rows(rows, mode="replace")
                self.selection_upload_info = f"조회조건 테이블에 {n}건을 자동 반영했습니다. (필요 시 직접 수정 가능)"
                return

            # target == 'alv'
            has_area = any("area" in r for r in rows)
            if has_area:
                a_rows = [r for r in rows if str(r.get("area") or "").strip().upper() != "B"]
                b_rows = [r for r in rows if str(r.get("area") or "").strip().upper() == "B"]
                na = self._apply_alv_rows(a_rows, area="A", mode="replace") if a_rows else 0
                nb = self._apply_alv_rows(b_rows, area="B", mode="replace") if b_rows else 0
                self._clear_se11_bindings_on_excel_fill()
                self.alv_upload_info = f"ALV 테이블 자동 반영 완료: 영역A {na}건, 영역B {nb}건"
            else:
                na = self._apply_alv_rows(rows, area="A", mode="replace")
                self._clear_se11_bindings_on_excel_fill()
                self.alv_upload_info = f"ALV(영역 A) 테이블에 {na}건을 자동 반영했습니다. (영역 B는 수동 또는 area 컬럼 사용)"
        except Exception as e:
            if target == "selection":
                self.selection_upload_error = f"엑셀 자동 입력 실패: {e}"
            else:
                self.alv_upload_error = f"엑셀 자동 입력 실패: {e}"

    @rx.event
    async def handle_excel_upload_selection(self, files: list[rx.UploadFile]):
        return await self._handle_excel_upload_target(files, target="selection")

    @rx.event
    async def handle_excel_upload_alv(self, files: list[rx.UploadFile]):
        return await self._handle_excel_upload_target(files, target="alv")

    # -----------------------------
    # SE11 이미지 → 업로드 테이블 목록 (이미지 1장 = 테이블 1개, 병합 없음)
    # -----------------------------
    def _data_uri_to_bytes(self, uri: str) -> bytes:
        u = (uri or "").strip()
        if not u.startswith("data:") or "," not in u:
            raise ValueError("유효한 이미지 data URI가 아닙니다.")
        header, b64 = u.split(",", 1)
        if ";base64" not in header:
            raise ValueError("지원하지 않는 이미지 인코딩입니다. (base64 필요)")
        return base64.b64decode(b64, validate=True)

    async def _extract_se11_table_from_image(self, data: bytes) -> tuple[str, List[ALVField]]:
        """이미지 1장에서 필드 목록 추출. 반환: (테이블명 힌트, ALV 행 목록)."""
        img = Image.open(io.BytesIO(data)).convert("RGB")
        api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
        genai.configure(api_key=api_key)
        model_name = (os.environ.get("GEMINI_MODEL") or "gemini-2.0-flash").strip()
        model = genai.GenerativeModel(model_name)
        prompt = (
            "다음 이미지는 SAP SE11(또는 DDIC) 테이블 구조 캡처입니다.\n"
            "JSON만 출력하세요. 다른 설명은 절대 포함하지 마세요.\n"
            "출력 형식:\n"
            "{\n"
            '  "table_name": "이미지에 보이는 테이블 이름(영문). 식별 불가면 빈 문자열",\n'
            '  "fields": [\n'
            '    {"field_id": "필드명", "label": "설명", "data_type": "타입", "length": "길이", "decimals": "소수", "key": true/false}\n'
            "  ]\n"
            "}\n"
            "필드명이 식별되지 않으면 해당 항목은 제외하고, 가능한 것만 넣으세요."
        )
        resp = model.generate_content([prompt, img])
        payload = self._extract_json(getattr(resp, "text", "") or "")
        name_hint = ""
        if isinstance(payload, dict):
            name_hint = self._clean_str(payload.get("table_name") or "")
        fields_raw = (payload.get("fields") if isinstance(payload, dict) else None) or []
        if not isinstance(fields_raw, list) or not fields_raw:
            raise ValueError("이미지에서 필드 정보를 추출하지 못했습니다.")

        alv_rows: List[ALVField] = []
        i = 0
        for item in fields_raw:
            if not isinstance(item, dict):
                continue
            fid = self._clean_str(item.get("field_id") or item.get("field") or item.get("name"))
            if not fid:
                continue
            lbl = self._clean_str(item.get("label") or item.get("description") or item.get("text"))
            is_key = self._to_bool(item.get("key") or item.get("is_key"))
            i += 1
            alv_rows.append(
                ALVField(
                    order=str(i),
                    label=lbl,
                    field_id=fid,
                    is_key=is_key,
                    is_sum=False,
                    is_edit=False,
                    action="",
                )
            )
        if not alv_rows:
            raise ValueError("필드명(field_id)을 추출하지 못했습니다.")
        return (name_hint, alv_rows)

    @rx.event
    def on_se11_image_paste(self, data: list[tuple[str, str]]):
        self.image_upload_error = ""
        added = 0
        for mime_type, item in data:
            if mime_type.startswith("image/"):
                self.se11_paste_queue_uris = [*self.se11_paste_queue_uris, item]
                added += 1
        if added:
            self.image_upload_info = ""
            return rx.toast(f"이미지 {added}장이 대기열에 추가되었습니다. 「테이블 생성」을 눌러주세요.")
        return rx.toast("클립보드에 이미지가 없습니다. 화면을 캡처한 뒤 이 영역에서 Ctrl+V 하세요.")

    @rx.event
    async def handle_se11_create_tables(self):
        """붙여넣은 이미지 각각을 개별 호출로 처리해 테이블 N개 생성(합치지 않음)."""
        self.image_upload_error = ""
        self.image_upload_info = ""

        api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
        if not api_key:
            self.image_upload_error = "API 키가 없습니다. `.env`에 GEMINI_API_KEY를 설정한 뒤 다시 시도해주세요."
            return

        jobs: list[bytes] = []
        for uri in self.se11_paste_queue_uris or []:
            u = (uri or "").strip()
            if not u:
                continue
            try:
                jobs.append(self._data_uri_to_bytes(u))
            except Exception as e:
                self.image_upload_error = f"붙여넣은 이미지를 읽지 못했습니다: {e}"
                return

        if not jobs:
            self.image_upload_error = "클립보드에 이미지를 붙여넣은 뒤 「테이블 생성」을 눌러주세요."
            return

        ok = 0
        err_msgs: list[str] = []
        for raw in jobs:
            try:
                name_hint, fields = await self._extract_se11_table_from_image(raw)
                label = self._unique_se11_table_label(name_hint if name_hint else "")
                ut = UploadedAlvTable(id=str(uuid.uuid4()), label=label, fields=fields)
                self.uploaded_alv_tables = [*self.uploaded_alv_tables, ut]
                ok += 1
            except Exception as e:
                err_msgs.append(str(e))

        self._sync_se11_select_options()
        self.se11_paste_queue_uris = []
        if ok:
            self.image_upload_info = f"테이블 {ok}개를 생성했습니다. 목록에서 영역 A/B에 바인딩하세요."
        if err_msgs:
            self.image_upload_error = "일부 이미지 처리 실패:\n" + "\n".join(err_msgs[:5])
            if len(err_msgs) > 5:
                self.image_upload_error += f"\n… 외 {len(err_msgs) - 5}건"

    def _read_fs_template(self) -> str:
        template_path = os.path.join("templates", "FS_Functional_Spec.md")
        if not os.path.exists(template_path):
            raise FileNotFoundError(
                f"FS 템플릿 파일을 찾을 수 없습니다: {template_path}"
            )
        with open(template_path, "r", encoding="utf-8") as f:
            return f.read()

    def _read_text_file(self, rel_path: str) -> str:
        if not os.path.exists(rel_path):
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {rel_path}")
        with open(rel_path, "r", encoding="utf-8") as f:
            return f.read()

    def _read_mapping_template(self) -> str:
        return self._read_text_file(os.path.join("templates", "FS_Mapping_Spec.md"))

    def _read_fs_to_mapping_rules(self) -> str:
        return self._read_text_file(os.path.join("rules", "fs_to_mapping_rules.md"))

    def _build_mapping_prompt(self, mapping_template: str, rules_text: str, fs_markdown: str) -> str:
        """
        NOTE:
        - rules_text 가 '생성 규칙(판단 기준)'의 단일 진실 소스.
        - mapping_template 는 출력 형식(섹션/표 구조) 유지 목적.
        """
        return f"""
당신은 SAP ABAP 개발 매핑 입력서 작성자입니다.

당신의 목표는 "FS(Functional Spec)"를 읽고, "FS → 개발 매핑 입력서 매핑 규칙(rules)"을 단 하나의 기준으로 삼아
개발 매핑 입력서(마크다운)를 생성하는 것입니다.

중요:
- "템플릿 기반으로 내용을 추측"하지 마세요.
- 판단/기본값/예외 처리는 아래 rules 내용을 그대로 적용하세요.
- 템플릿은 출력 형식(섹션/표/헤더/번호/컬럼)을 맞추기 위한 용도일 뿐, 생성 로직의 기준이 아닙니다.

출력 규칙:
- 출력은 마크다운만 반환
- 템플릿의 섹션/표 구조를 최대한 유지
- rules에 의해 결정 불가/FS에 근거 없음인 값은 rules의 기본값/예외를 적용하고, 그래도 없으면 'TBD'로 표기
- rules에 '미입력 시 오류'로 명시된 항목이 비어있다면, 해당 항목은 빈칸 대신 'ERROR: ...' 형태로 명확히 표시

아래는 출력 형식 템플릿(형식만 참고):
```md
{mapping_template}
```

아래는 생성 규칙(rules) (이 내용이 판단 기준):
```md
{rules_text}
```

아래는 입력 FS 원문:
```md
{fs_markdown}
```
""".strip()

    @staticmethod
    def _split_md_sections(md: str) -> dict[str, str]:
        """
        Split markdown by '## ' headings.
        Returns {heading_title: block_including_heading_line}.
        """
        text = (md or "").replace("\r\n", "\n")
        parts = re.split(r"(?m)^##\s+", text)
        out: dict[str, str] = {}
        if not parts:
            return out
        # parts[0] is preface (e.g. '# ...')
        for p in parts[1:]:
            lines = p.split("\n")
            title = (lines[0] or "").strip()
            body = "\n".join(lines[1:]).strip("\n")
            out[title] = f"## {title}\n{body}".strip() + "\n"
        return out

    @staticmethod
    def _tbd_signal(md: str) -> dict[str, bool]:
        """
        Coarse TBD propagation signal by FS section headings.
        """
        sections = State._split_md_sections(md)

        def has_tbd(title_contains: str) -> bool:
            for title, block in sections.items():
                if title_contains in title and "TBD" in block:
                    return True
            return False

        return {
            "FS5": has_tbd("5."),
            "FS6": has_tbd("6."),
            "FS8": has_tbd("8."),
            "FS9_1": has_tbd("9.1") or has_tbd("9. 1") or has_tbd("9.1"),
            "FS9_2": has_tbd("9.2") or has_tbd("9. 2") or has_tbd("9.2"),
            "FS10": has_tbd("10."),
        }

    def _build_mapping_sections_prompt(self, template_md: str, rules_md: str, fs_md: str) -> str:
        template_sections = self._split_md_sections(template_md)

        # Provide only relevant template blocks to guide formatting per section.
        def tpl(title_contains: str) -> str:
            for title, block in template_sections.items():
                if title_contains in title:
                    return block.strip()
            return ""

        tbd = self._tbd_signal(fs_md)

        return f"""
당신은 SAP ABAP 개발 매핑 입력서 작성자입니다.

목표:
- Functional Spec(FS)와 rules를 기준으로 개발 매핑 입력서를 생성합니다.
- 개발 매핑 입력서는 '섹션 단위'로 독립 생성합니다.
- 사용자는 생성 후 섹션별로 편집할 수 있으므로, 각 섹션은 완결된 마크다운 블록이어야 합니다.

중요(이번 작업의 핵심):
- 생성 판단 기준은 오직 rules 문서입니다. 템플릿은 출력 형식 유지 용도입니다.
- FS 내 정보가 부족해 TBD가 있는 경우, 그에 의존하는 매핑 섹션도 동일하게 TBD로 유지하세요.
  - FS 5 TBD → ALV 매핑 섹션을 TBD 중심으로 생성
  - FS 6 TBD → 저장 흐름 / FORM 정의 섹션을 TBD 중심으로 생성
  - FS 8 TBD → 기능 매핑 / FORM 정의 / 이벤트 매핑 섹션을 TBD 중심으로 생성

생성 기준(섹션별 FS 기준):
- 프로그램 기본 정보: FS 1,2,3,5,6
- Include 구성: FS 전체
- 기능 매핑: FS 8
- FORM 정의: FS 8
- 조회 DB 매핑: FS 9.1
- 저장 DB 매핑: FS 9.2
- ALV 매핑: FS 5
- 이벤트 매핑: FS 8
- 저장 흐름: FS 6
- 메시지: FS 10

출력 형식:
- 아래 JSON만 반환하세요. 다른 텍스트/설명 금지.
- 각 값은 마크다운 문자열이며, 템플릿의 표 구조/컬럼을 최대한 유지하세요.

JSON 스키마:
{{
  "프로그램 기본 정보": "<md>",
  "Include 구성": "<md>",
  "기능 매핑": "<md>",
  "FORM 정의": "<md>",
  "조회 DB 매핑": "<md>",
  "저장 DB 매핑": "<md>",
  "ALV 매핑": "<md>",
  "이벤트 매핑": "<md>",
  "저장 흐름": "<md>",
  "메시지": "<md>"
}}

템플릿(형식 가이드):
- 프로그램 기본 정보:
```md
{tpl("1. 프로그램 기본 정보")}
```
- Include 구성:
```md
{tpl("2. Include 구성")}
```
- 기능 매핑:
```md
{tpl("3. 기능 매핑")}
```
- FORM 정의:
```md
{tpl("4. FORM 정의")}
```
- 조회 DB 매핑:
```md
{tpl("5. 조회 DB 매핑")}
```
- 저장 DB 매핑:
```md
{tpl("6. 저장 DB 매핑")}
```
- ALV 매핑:
```md
{tpl("7. ALV 매핑")}
```
- 이벤트 매핑:
```md
{tpl("8. 이벤트 매핑")}
```
- 저장 흐름:
```md
{tpl("9. 저장 흐름")}
```
- 메시지:
```md
{tpl("10. 메시지")}
```

rules(판단 기준 단일 소스):
```md
{rules_md}
```

FS 원문:
```md
{fs_md}
```

TBD 신호(참고): {json.dumps(tbd, ensure_ascii=False)}
""".strip()

    @rx.event
    def set_mapping_active_section(self, value: str):
        self.mapping_active_section = value

    @rx.event
    def set_mapping_edit_section(self, value: str):
        """개발 매핑 입력서 섹션 편집 모드(단일 섹션) 토글용."""
        self.mapping_edit_section = value or ""

    @rx.event
    def update_mapping_section(self, section: str, value: str):
        curr = dict(self.mapping_sections or {})
        curr[section] = value
        self.mapping_sections = curr
        self.mapping_updated_at = datetime.now(timezone.utc).isoformat()

    def _ensure_mapping_sections(self):
        if self.mapping_sections:
            return
        self.mapping_sections = {
            "프로그램 기본 정보": "",
            "Include 구성": "",
            "기능 매핑": "",
            "FORM 정의": "",
            "조회 DB 매핑": "",
            "저장 DB 매핑": "",
            "ALV 매핑": "",
            "이벤트 매핑": "",
            "저장 흐름": "",
            "메시지": "",
        }

    def _apply_mapping_sections_payload(self, payload: dict):
        self._ensure_mapping_sections()
        curr = dict(self.mapping_sections or {})

        def strip_leading_headings(md: str) -> str:
            """Remove only the first leading section heading to avoid duplicate titles in UI.

            Keep sub-headings like '### ...' (e.g., '### FS 전체 구조 기반 자동 결정').
            """
            lines = (md or "").replace("\r\n", "\n").split("\n")
            i = 0
            # Skip leading blank lines
            while i < len(lines) and lines[i].strip() == "":
                i += 1
            # Drop only the first top heading line (usually "## <section title>" or "# ...")
            if i < len(lines) and lines[i].lstrip().startswith("#"):
                i += 1
                while i < len(lines) and lines[i].strip() == "":
                    i += 1
            return "\n".join(lines[i:]).strip()

        for k in list(curr.keys()):
            v = payload.get(k)
            if isinstance(v, str):
                curr[k] = strip_leading_headings(v)
        self.mapping_sections = curr
        self.mapping_updated_at = datetime.now(timezone.utc).isoformat()

    def generate_mapping_sections(self):
        """
        FS(self.generated_fs) 기반으로 rules를 적용해 섹션 단위로 개발 매핑 입력서를 생성.
        """
        self.mapping_error = ""
        self.generated_mapping_spec = ""
        self._ensure_mapping_sections()

        api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
        if not api_key:
            self.mapping_error = "API 키가 없습니다. `.env`에 GEMINI_API_KEY를 설정해주세요."
            return
        fs_md = (self.generated_fs or "").strip()
        if not fs_md:
            self.mapping_error = "FS가 비어 있어 개발 매핑 입력서를 생성할 수 없습니다. 먼저 FS를 생성하세요."
            return

        try:
            template_md = self._read_mapping_template()
            rules_md = self._read_fs_to_mapping_rules()
            prompt = self._build_mapping_sections_prompt(template_md, rules_md, fs_md)

            genai.configure(api_key=api_key)
            preferred = (os.environ.get("GEMINI_MODEL") or "").strip()
            candidates = [m for m in [preferred, "gemini-2.0-flash", "gemini-1.5-pro-latest", "gemini-1.5-flash-latest"] if m]

            last_err: Exception | None = None
            resp = None
            for name in candidates:
                try:
                    resp = genai.GenerativeModel(name).generate_content(prompt)
                    last_err = None
                    break
                except Exception as e:
                    last_err = e

            if resp is None:
                raise RuntimeError(f"개발 매핑 입력서(섹션) 생성 모델 호출 실패: {last_err}")

            payload = self._extract_json(getattr(resp, "text", "") or "")
            if not isinstance(payload, dict):
                raise ValueError("매핑 섹션 생성 결과(JSON)가 dict 형태가 아닙니다.")
            self._apply_mapping_sections_payload(payload)

            # Also keep a single combined doc (legacy)
            combined = []
            for sec in [
                "프로그램 기본 정보",
                "Include 구성",
                "기능 매핑",
                "FORM 정의",
                "조회 DB 매핑",
                "저장 DB 매핑",
                "ALV 매핑",
                "이벤트 매핑",
                "저장 흐름",
                "메시지",
            ]:
                block = (self.mapping_sections.get(sec) or "").strip()
                if block:
                    combined.append(block)
            self.generated_mapping_spec = "\n\n---\n\n".join(combined).strip()
        except Exception as e:
            self.mapping_error = str(e)

    def generate_mapping_spec(self):
        """
        FS 결과(self.generated_fs)를 기반으로 rules에 따라 개발 매핑 입력서를 생성.
        """
        self.mapping_error = ""
        self.generated_mapping_spec = ""

        api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
        if not api_key:
            self.mapping_error = "API 키가 없습니다. `.env`에 GEMINI_API_KEY를 설정해주세요."
            return
        if not (self.generated_fs or "").strip():
            self.mapping_error = "FS가 비어 있어 개발 매핑 입력서를 생성할 수 없습니다. 먼저 FS를 생성하세요."
            return

        try:
            mapping_template = self._read_mapping_template()
            rules_text = self._read_fs_to_mapping_rules()
            prompt = self._build_mapping_prompt(mapping_template, rules_text, self.generated_fs)

            genai.configure(api_key=api_key)
            preferred = (os.environ.get("GEMINI_MODEL") or "").strip()
            candidates = [m for m in [preferred, "gemini-2.0-flash", "gemini-1.5-pro-latest", "gemini-1.5-flash-latest"] if m]

            last_err: Exception | None = None
            resp = None
            for name in candidates:
                try:
                    resp = genai.GenerativeModel(name).generate_content(prompt)
                    last_err = None
                    break
                except Exception as e:
                    last_err = e

            if resp is None:
                raise RuntimeError(f"개발 매핑 입력서 생성 모델 호출 실패: {last_err}")

            self.generated_mapping_spec = (resp.text or "").strip()
            if not self.generated_mapping_spec:
                raise RuntimeError("개발 매핑 입력서 생성 결과가 비어 있습니다.")
        except Exception as e:
            self.mapping_error = str(e)

    def _build_fs_prompt(self, template: str) -> str:
        selection_rows = [
            {
                "order": f.order,
                "section": f.section,
                "label": f.label,
                "field_id": f.field_id,
                "input_type": f.input_type,
                "is_required": f.is_required,
                "has_f4": f.has_f4,
                "ui_rule": f.ui_rule,
            }
            for f in self.selection_fields
        ]
        # 섹션 5.2 "결과 필드 정의"는 사용자가 체크한 "선택한 필드 요약"을 우선 사용.
        # (체크가 하나도 없을 때만 현재 편집 표 전체를 fallback으로 사용)
        alv_a_src = self.alv_picked_preview_all_a if (self.alv_picked_preview_all_a or []) else self.alv_fields_a
        alv_b_src = self.alv_picked_preview_all_b if (self.alv_picked_preview_all_b or []) else self.alv_fields_b

        alv_a = [
            {
                "order": f.order,
                "label": f.label,
                "field_id": f.field_id,
                "is_key": f.is_key,
                "is_sum": f.is_sum,
                "is_edit": f.is_edit,
                "action": f.action,
            }
            for f in alv_a_src
        ]
        alv_b = [
            {
                "order": f.order,
                "label": f.label,
                "field_id": f.field_id,
                "is_key": f.is_key,
                "is_sum": f.is_sum,
                "is_edit": f.is_edit,
                "action": f.action,
            }
            for f in alv_b_src
        ]

        # 템플릿 구조를 유지하면서, 아래 입력값으로 빈 칸을 채우도록 유도
        return f"""
당신은 SAP ABAP Functional Spec 작성자입니다.
아래 "템플릿"의 구조(헤더/섹션/표/번호)를 최대한 유지하면서, "입력 데이터"를 사용해 내용을 채워서
완성된 Functional Spec 문서를 한국어로 작성하세요.

규칙:
- 출력은 마크다운만 반환
- 템플릿의 섹션 순서/형식을 유지 (추가 섹션이 필요하면 맨 아래에 최소로)
- 값이 비어있는 항목은 "TBD"로 표기
- 표 형태가 있는 템플릿은 표를 유지

템플릿:
```md
{template}
```

입력 데이터(JSON):
```json
{{
  "program_basic": {{
    "prog_name": "{self.prog_name}",
    "req_dept_user": "{self.req_dept_user}",
    "work_category": "{self.work_category}",
    "ref_tcode": "{self.ref_tcode}",
    "auth_range": {self.auth_range}
  }},
  "program_type": {{
    "prog_type": "{self.prog_type}",
    "selected_features": {self.selected_features}
  }},
  "ui_layout": {{
    "ui_layout": "{self.ui_layout}",
    "layout_detail": "{self.layout_detail}",
    "ui_features": {self.ui_features},
    "edit_auth": "{self.edit_auth}",
    "other_ui_notes": "{self.other_ui_notes}"
  }},
  "selection_screen": {selection_rows},
  "alv_area_a": {alv_a},
  "alv_area_b": {alv_b},
  "business_rules": {json.dumps(
        {
            "logic_flow": self.logic_flow,
            "validation_rule": self.validation_rule,
            "auto_calc_rule": self.auto_calc_rule,
            "data_target": self.data_target,
            "delete_method": self.delete_method,
            "dup_check": self.dup_check,
        },
        ensure_ascii=False,
    )}
}}
```
""".strip()

    def generate_fs(self):
        self.is_generating_fs = True
        self.fs_error = ""
        self.generated_fs = ""
        self.mapping_error = ""
        self.generated_mapping_spec = ""

        try:
            api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
            if not api_key:
                raise RuntimeError(
                    "GEMINI_API_KEY(또는 GOOGLE_API_KEY)가 없습니다. "
                    "Windows 환경 변수에 설정하거나, 프로젝트 루트의 .env 파일에 넣어 주세요."
                )

            template = self._read_fs_template()
            prompt = self._build_fs_prompt(template)

            genai.configure(api_key=api_key)
            # Model selection:
            # - Prefer explicit env override
            # - Otherwise try common stable aliases
            # - If those fail (404 / unsupported), pick the first model that supports generateContent.
            preferred = (os.environ.get("GEMINI_MODEL") or "").strip()
            candidates = [m for m in [preferred, "gemini-1.5-pro-latest", "gemini-2.0-flash", "gemini-1.5-flash-latest"] if m]

            last_err: Exception | None = None
            resp = None
            for name in candidates:
                try:
                    resp = genai.GenerativeModel(name).generate_content(prompt)
                    last_err = None
                    break
                except Exception as e:
                    last_err = e

            if resp is None:
                try:
                    # Fallback: discover an available model supporting generateContent.
                    for m in genai.list_models():
                        methods = getattr(m, "supported_generation_methods", []) or []
                        if "generateContent" in methods or "generate_content" in methods:
                            model_name = getattr(m, "name", "")
                            if model_name.startswith("models/"):
                                model_name = model_name.replace("models/", "", 1)
                            resp = genai.GenerativeModel(model_name).generate_content(prompt)
                            last_err = None
                            break
                except Exception as e:
                    last_err = e

            if resp is None:
                raise RuntimeError(f"사용 가능한 Gemini 모델을 찾지 못했습니다. 마지막 오류: {last_err}")

            self.generated_fs = (resp.text or "").strip()
            self._refresh_fs_blocks()

            # 생성 후 FS 그룹으로 열어두기
            self.expanded_sidebar_group = "fs"
            self.current_section = self.fs_sections[0]

            if not self.generated_fs:
                raise RuntimeError("FS 생성 결과가 비어 있습니다. (Gemini 응답이 empty)")

            # FS 생성 성공 시, rules 기반으로 개발 매핑 입력서도 함께 생성
            self.generate_mapping_sections()
            return rx.window_alert("✅ FS 생성 완료! (FS 탭에서 확인하세요)")
        except Exception as e:
            self.fs_error = str(e)
            return rx.window_alert(f"❌ FS 생성 실패: {e}")
        finally:
            self.is_generating_fs = False

    # state.py 에 이 함수가 있는지 확인이 필요합니다

    def set_ref_tcode(self, value: str):
        self.ref_tcode = value

    def update_data_target(self, value: str, checked: bool):
        if checked:
            self.data_target.append(value)
        else:
            self.data_target = [i for i in self.data_target if i != value]

    # -----------------------------
    # 입력 보완(팝업) → 원본 데이터 수정
    # -----------------------------
    @rx.event
    def open_edit_modal(self):
        self.edit_modal_open = True
        self.edit_modal_error = ""
        # preload current value
        self.edit_field_value = str(getattr(self, self.edit_field_key, "") or "")

    @rx.event
    def close_edit_modal(self):
        self.edit_modal_open = False
        self.edit_modal_error = ""

    @rx.event
    def set_edit_field_key(self, value: str):
        self.edit_field_key = value
        self.edit_field_value = str(getattr(self, value, "") or "")

    @rx.event
    def set_edit_field_value(self, value: str):
        self.edit_field_value = value

    @rx.event
    def apply_edit_and_regenerate(self):
        """
        원본 데이터(Single Source of Truth)를 수정한 뒤 FS/매핑 재생성.
        """
        self.edit_modal_error = ""
        key = (self.edit_field_key or "").strip()
        if not key or not hasattr(self, key):
            self.edit_modal_error = "수정할 항목이 올바르지 않습니다."
            return

        # Apply
        setattr(self, key, self.edit_field_value)
        self.edit_modal_open = False

        # Regenerate FS + Mapping (FS is read-only output)
        return self.generate_fs()
